import os
import pandas as pd
import openpyxl
from pathlib import Path
import customtkinter as ctk
from tela import Tela

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

def executar_automacao():
    arquivo_orignal = tela.entrada_origem.get().strip()
    pasta_final = tela.entrada_destino.get().strip()

    if not arquivo_orignal or not pasta_final:
        tela.label_status.configure(text="Erro: Preencha a origem e o destino!", text_color="red")
        return

    tela.label_status.configure(text="Processando. Por favor, aguarde.", text_color="yellow")

    # Força a interface a atualizar o texto antes de travar no processo pesado
    app.update_idletasks()
    
    arquivo_novo = Path(pasta_final) / "Alocação por Porto.xlsx"
    aloc = {}

    config_tipos = {
        "Cód. Porto": str,
    }

    abas_criadas = []

    df_alocacao = pd.read_excel(arquivo_orignal, sheet_name='Alocação', header=0, index_col=0, dtype=config_tipos)

    df_base = pd.read_excel(arquivo_orignal, sheet_name='Base', header=0, index_col=0, dtype=config_tipos)

    wb = openpyxl.load_workbook(arquivo_orignal, data_only=False)

    aba_base = wb['Base']
    cabecalhos = [cell.value for cell in aba_base[1]]

    for index, data_aloc in df_alocacao.iterrows():
        aloc[data_aloc['Cód. Embarcação']] = data_aloc['Cód. Porto']

    for index, data_aloc in df_alocacao.iterrows():
        nome_aba = data_aloc['Cód. Porto'][:10]

        if nome_aba not in wb.sheetnames:
            wb.create_sheet(title=nome_aba)
            aba = wb[nome_aba]
            aba.append(cabecalhos)
            print(f"Aba '{nome_aba}' criada com sucesso!")
            abas_criadas.append(nome_aba)

    for index, data_base in df_base.iterrows():
        if data_base['Cód. Porto'] != aloc[data_base['Cód. Embarcação']]:
            if data_base['Cód. Porto'] in abas_criadas:
                aba = wb[aloc[data_base['Cód. Embarcação']]]
                aba.append(data_base.tolist())
                print(f"Dados da linha {index} adicionados à aba '{nome_aba}'.")

    wb.save(arquivo_novo)

    tela.label_status.configure(text="Concluído! O arquivo 'Alocação por Porto.xlsx' foi gerado com sucesso!", text_color="green")

app = ctk.CTk()
tela = Tela(app, executar_automacao)

app.mainloop()

